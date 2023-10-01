# openpyxl https://openpyxl.readthedocs.io/en/stable/tutorial.html#installation
import openpyxl as op

# 새로운 Workbook 객체 생성
wb = op.Workbook()

# print(wb)

# .save : 파일 저장
wb.save("test.xlsx")


# .create_sheet : 엑셀 아래에 시트 생성
# .load_workbook : 워크북 객체 생성
# .active 활성화 되어있는 시트 설정

# .cell

# 방법 1 : Sheet의 Cell 속성 사용하기
# data1 = ws.cell(row=1, column=2).value
#
# #방법 2 : 엑셀 인덱스(Range) 사용하기
# data2 = ws["B1"].value
#
# #위 결과 출력해보기
# print("cell(1,2) : ", data1)
# print('Range("B1"):', data2)
